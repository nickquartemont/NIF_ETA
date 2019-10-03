"""
Fission Product Prediction Tool 
Utilizes General Observables of Fission (GEF) and experimental data fit to a 
phenomenological function (Nagy Fit Y=Y_0*exp(-b*E)) to predict excpected 
fission products from Pu-239, Pu-240, U-235, U-236, and U-238 neutron induced fission.
 
GEF data is utilized to determine the entire mass chain. 
Nagy fit data is utilized for individual cumulative yields of selected isotopes 
when available for all fissioning systems. 
Additional isotopes can be added to the excel files by using the same 
convention. 

See White paper and Powerpoint for more information. 

"""

# In order to work as an executable, the functions needed to be in here. 

import os 
import numpy as np
import sys
import datetime
from math import sqrt
from collections import defaultdict
import openpyxl
from math import exp
from scipy.optimize import curve_fit
from scipy import interpolate
import time
from scipy.interpolate import interp1d
import pickle 

#---------------------------------------------------------------------------------------#  

def GatherInput(CWD,file,Files):
    """
    Check files that are requested from input folder. This 
    search checks for U-235/8 and Pu-239/40 input spectra. 
    """
    # Initialize Data 
   # Import Data 
    Pu239=[0.0]
    Pu240=[0.0]
    U238=[0.0]
    U235=[0.0]
    U236=[0.0]
    Flag2=np.zeros(5)
    # Modifying energy values to be bin midpoints 
    for f in Files: 
        if f=='Pu239.txt': 
            Pu239=np.loadtxt(open(CWD+'\\Inputs\\'+f, "rb"), delimiter=",", skiprows=0)
            Pu239=Echeck(Pu239)
            Flag2[0]=Echeck3(Pu239,CWD)
        if f=='Pu240.txt': 
            Pu240=np.loadtxt(open(CWD+'\\Inputs\\'+f, "rb"), delimiter=",", skiprows=0)
            Pu240=Echeck(Pu240)
            Flag2[1]=Echeck3(Pu240,CWD)
        if f=='U235.txt': 
            U235=np.loadtxt(open(CWD+'\\Inputs\\'+f, "rb"), delimiter=",", skiprows=0)
            U235=Echeck(U235)
            Flag2[2]=Echeck3(U235,CWD)
        if f=='U238.txt': 
            U238=np.loadtxt(open(CWD+'\\Inputs\\'+f, "rb"), delimiter=",", skiprows=0)
            U238=Echeck(U238)
            Flag2[3]=Echeck3(U238,CWD)
        if f=='U236.txt': 
            U236=np.loadtxt(open(CWD+'\\Inputs\\'+f, "rb"), delimiter=",", skiprows=0)
            U236=Echeck(U236)
            Flag2[4]=Echeck3(U236,CWD)
    Track=np.zeros(5)
    if len(U235)>1: 
        Track[0]=1
    if len(U236)>1: 
        Track[1]=1
    if len(U238)>1: 
        Track[2]=1
    if len(Pu239)>1: 
        Track[3]=1
    if len(Pu240)>1: 
        Track[4]=1
    if np.sum(Track)==np.sum(Flag2):
        Flag2=0
    else: 
        Flag2=1

    Fissioning={}
    Fissioning['U235']=U235
    Fissioning['U236']=U236
    Fissioning['U238']=U238
    Fissioning['Pu239']=Pu239
    Fissioning['Pu240']=Pu240
    Fissioning['Track']=Track
    # Get weights 
    Weights=np.zeros(5)
    if os.path.isfile(CWD+'\\Inputs\\'+'Weights.txt')==True: 
        Weights=np.loadtxt(open(CWD+'\\Inputs\\'+'Weights.txt', "rb"), delimiter=",", skiprows=0,usecols=1)
    else: 
        if Fissioning['Track'][0]==1:
            Weights[0]=np.sum(Fissioning['U235'][:,1])
        if Fissioning['Track'][1]==1:
            Weights[1]=np.sum(Fissioning['U236'][:,1])
        if Fissioning['Track'][2]==1:
            Weights[2]=np.sum(Fissioning['U238'][:,1])
        if Fissioning['Track'][3]==1:
            Weights[3]=np.sum(Fissioning['Pu239'][:,1])
        if Fissioning['Track'][4]==1:
            Weights[4]=np.sum(Fissioning['Pu240'][:,1])
    Weights=Weights/np.sum(Weights)
    if len(Weights)!=5: 
        file.write("Error: Must include weights for each system. Set non-used sytems to zero")
    Weights=Weights/np.sum(Weights)
    file.write("Using weights for the total fissions in each system of\n")
    file.write("U-235 = {:04.3f}\n".format(Weights[0]))
    file.write("U-236 = {:04.3f}\n".format(Weights[1]))
    file.write("U-238 = {:04.3f}\n".format(Weights[2]))
    file.write("Pu-239 = {:04.3f}\n".format(Weights[3]))
    file.write("Pu-240 = {:04.3f}\n".format(Weights[4]))
    Fissioning["Weights"]=Weights
    return file, Fissioning,Flag2,Track

def Echeck(Pu239):
    if Pu239[0,0]>Pu239[-1,0]: 
        file.write('Energy needs to be monotonically increasing')
    Pu239[1:,0]=Pu239[1:,0]-(Pu239[1:,0]-Pu239[0:-1,0])/2.0
    Pu239[0,0]=Pu239[0,0]/2.0
    return Pu239
def Echeck2(Fissioning,file): 
    Sizes=[]
    Flag=0
    if Fissioning['Track'][0]==1: 
        Sizes.append(len(Fissioning['U235']))
        Energy=Fissioning['U235'][:,0]
    if Fissioning['Track'][1]==1: 
        Sizes.append(len(Fissioning['U236']))
        Energy=Fissioning['U236'][:,0]
    if Fissioning['Track'][2]==1: 
        Sizes.append(len(Fissioning['U238']))
        Energy=Fissioning['U238'][:,0]
    if Fissioning['Track'][3]==1: 
        Sizes.append(len(Fissioning['Pu239']))
        Energy=Fissioning['Pu239'][:,0]
    if Fissioning['Track'][4]==1: 
        Sizes.append(len(Fissioning['Pu240']))
        Energy=Fissioning['Pu240'][:,0]
    if np.max(Sizes)!=np.min(Sizes):
        file.write("Error in results\n")
        file.write("Energy groups should be consistant amongst input spectra\n")
        Flag=1
    return Flag,Energy

def Echeck3(fissioning,CWD): 
    Dplus=np.loadtxt(open(CWD+"\\Data\\E_Bins.csv", "rb"), delimiter=",", skiprows=1)
    Edata=Dplus[:,2]
    Edata2=fissioning[:,0]
    Flag2=0
    if len(Edata)==len(Edata2):
        # At least the group size is the same
        Diff=np.divide(np.subtract(Edata,Edata2).reshape(46,1),Edata.reshape(46,1))
        if np.max(np.abs(Diff))<0.005: # These look the same
            Flag2=1
    return Flag2

def run_GEF(Fissioning,file,Files,Flag2,CWD): 
    Flag=0
    Flag,E1=Echeck2(Fissioning,file)
    file.write("\n")
    if Flag2==1: 
        file.write("Calculating GEF results with interpolation based on group structure provided\n")
    if Flag2==0: 
        file.write("Calculating GEF results with DPLUS 46 Group Structure\n")
    if Flag==0:    
        if Fissioning['Track'][0]==1: 
            Datafiles=open(CWD+"\\Data\\u235_GEF.pckl",'rb')
            Gef235=pickle.load(Datafiles)
            Datafiles.close()
            if Flag2==0:         # Check if using DPLUS        
                y235,e235=GEF_FitDPLUS(Gef235,Fissioning['U235'],Fissioning['Weights'][0])
            if Flag2==1:         # If not use other group 
                for i in range(155,181):
                    Gef235.pop(str(i))            
                for i in range(56,58):
                    Gef235.pop(str(i))            
                for i in range(59,78):
                    Gef235.pop(str(i))
                y235,e235=GEF_Fit(Gef235,Fissioning,Fissioning['Weights'][0],'U235',CWD)
            file.write("Finished U-235 (n,f) GEF results\n")
            print("Finished U-235 (n,f) GEF results")    
            file.write("")
    
        if Fissioning['Track'][1]==1: 
            Datafiles=open(CWD+"\\Data\\u236_GEF.pckl",'rb')
            Gef236=pickle.load(Datafiles)
            Datafiles.close()
            if Flag2==0:         # Check if using DPLUS 
                y236,e236=GEF_FitDPLUS(Gef236,Fissioning['U236'],Fissioning['Weights'][1])
            if Flag2==1:         # If not use other group 
                for i in range(155,181):
                    Gef236.pop(str(i))            
                for i in range(56,58):
                    Gef236.pop(str(i))            
                for i in range(59,78):
                    Gef236.pop(str(i))
                y236,e236=GEF_Fit(Gef236,Fissioning,Fissioning['Weights'][1],'U236',CWD)
            file.write("Finished U-236 (n,f) GEF results\n")
            print("Finished U-236 (n,f) GEF results")
            file.write("")
    
        if Fissioning['Track'][2]==1: 
            Datafiles=open(CWD+"\\Data\\u238_GEF.pckl",'rb')
            Gef238=pickle.load(Datafiles)
            Datafiles.close()
            if Flag2==0:         # Check if using DPLUS 
                y238,e238=GEF_FitDPLUS(Gef238,Fissioning['U238'],Fissioning['Weights'][2])
            if Flag2==1:         # If not use other group 
                for i in range(155,181):
                    Gef238.pop(str(i))            
                for i in range(56,58):
                    Gef238.pop(str(i))            
                for i in range(59,78):
                    Gef238.pop(str(i))
                y238,e238=GEF_Fit(Gef238,Fissioning,Fissioning['Weights'][2],'U238',CWD)
            file.write("Finished U-238 (n,f) GEF results\n")
            print("Finished U-238 (n,f) GEF results")
            file.write("")
    
        if Fissioning['Track'][3]==1: 
            Datafiles=open(CWD+"\\Data\\pu239_GEF.pckl",'rb')
            Gef239=pickle.load(Datafiles)
            Datafiles.close()
            if Flag2==0:         # Check if using DPLUS 
                print 'here'
                y239,e239=GEF_FitDPLUS(Gef239,Fissioning['Pu239'],Fissioning['Weights'][3])
            if Flag2==1:         # If not use other group 
                for i in range(155,181):
                    Gef239.pop(str(i))            
                for i in range(56,58):
                    Gef239.pop(str(i))            
                for i in range(59,78):
                    Gef239.pop(str(i))
                y239,e239=GEF_Fit(Gef239,Fissioning,Fissioning['Weights'][3],'Pu239',CWD)
            file.write("Finished Pu-239 (n,f) GEF results\n")
            print("Finished Pu-239 (n,f) GEF results")
        
        file.write("")
        if Fissioning['Track'][4]==1: 
            Datafiles=open(CWD+"\\Data\\pu240_GEF.pckl",'rb')
            Gef240=pickle.load(Datafiles)
            Datafiles.close()
            if Flag2==0:         # Check if using DPLUS 
                y240,e240=GEF_FitDPLUS(Gef240,Fissioning['Pu240'],Fissioning['Weights'][4])
            if Flag2==1:         # If not use other group 
                for i in range(155,181):
                    Gef240.pop(str(i))            
                for i in range(56,58):
                    Gef240.pop(str(i))            
                for i in range(59,78):
                    Gef240.pop(str(i))
                y240,e240=GEF_Fit(Gef240,Fissioning,Fissioning['Weights'][4],'Pu240',CWD)
            file.write("Finished Pu-240 (n,f) GEF results\n")
            print("Finished Pu-240 (n,f) GEF results")
    
    print ("collecting results")
    FirstTrack=Fissioning['Track'].tolist().index(filter(lambda x: x!=0, Fissioning['Track'].tolist())[0])
    y={}
    err={}
    if FirstTrack==0: 
        for k in y235.keys():
            y[k]=0.0
            err[k]=0.0        
    if FirstTrack==1: 
        for k in y236.keys():
            y[k]=0.0
            err[k]=0.0   
    if FirstTrack==2: 
        for k in y238.keys():
            y[k]=0.0
            err[k]=0.0  
    if FirstTrack==3: 
        for k in y239.keys():
            y[k]=0.0
            err[k]=0.0  
    if FirstTrack==4: 
        for k in y240.keys():
            y[k]=0.0
            err[k]=0.0           
    if Fissioning['Track'][0]==1: 
        for k in y235.keys(): 
            y[k]=y[k]+y235[k]
            err[k]=np.sqrt(err[k]**2+e235[k]**2)
    if Fissioning['Track'][1]==1: 
        for k in y236.keys(): 
            y[k]=y[k]+y236[k]
            err[k]=np.sqrt(err[k]**2+e236[k]**2)
    if Fissioning['Track'][2]==1: 
        for k in y238.keys(): 
            y[k]=y[k]+y238[k]
            err[k]=np.sqrt(err[k]**2+e238[k]**2)
    if Fissioning['Track'][3]==1: 
        for k in y239.keys(): 
            y[k]=y[k]+y239[k]
            err[k]=np.sqrt(err[k]**2+e239[k]**2)
    if Fissioning['Track'][4]==1: 
        for k in y240.keys(): 
            y[k]=y[k]+y240[k]
            err[k]=np.sqrt(err[k]**2+e240[k]**2)
    return file,y,err

def GEF_Fit(Gef235,Fissioning,Fis_Weight,Sys,CWD):
    y={}
    err={}
    err_fit={}
    err_flux={}
    err_data={}
    for k in Gef235.keys():
        Dplus=np.loadtxt(open(CWD+"\\Data\\E_Bins.csv", "rb"), delimiter=",", skiprows=1)
        Edata=Fissioning[Sys][:,0]
        f=interp1d(Dplus[:,2],Gef235[k][:,1],fill_value='extrapolate')
        Uncert=f(Edata)
        A=k
        y[A]=0.0
        err[A]=0.0
        err_fit[A] = 0.0
        err_flux[A] = 0.0
        err_data[A] = 0.0
        # Determine fitting parameters
        if A<=106 or A>=131:
            try: 
                params1,covar1=curve_fit(Nagy, Edata, Gef235[k][:,0],sigma=Gef235[k][:,1],absolute_sigma=True)
                params2=params1
                perr1 = np.sqrt(np.diag(covar1))
                perr2 = perr1
            except RuntimeError: # This was added for values that are all zero. 
               params1=np.zeros(2)
               params2=params1
               perr1=params1
               perr2=params1
               covar1=np.zeros([2,2])
               covar2=covar1
        else:
            ind=0
            while Edata[ind]<=5.5:
                ind+=1
            try: 
                params1,covar1=curve_fit(Nagy, Edata[0:ind], Gef235[k][0:ind,0],sigma=Gef235[k][0:ind,1],absolute_sigma=True)
            except RuntimeError: # This was added for values that are all zero. 
               params1=np.zeros(2)
               params2=params1
               perr1=params1
               perr2=params1
               covar1=np.zeros([2,2])
               covar2=covar1
            try: 
                params2,covar2=curve_fit(Nagy, Edata[ind:], Gef235[k][ind:,0],sigma=Gef235[k][ind:,1],absolute_sigma=True)
            except RuntimeError: # This was added for values that are all zero. 
               params1=np.zeros(2)
               params2=params1
               perr1=params1
               perr2=params1
               covar1=np.zeros([2,2])
               covar2=covar1
            perr1 = np.sqrt(np.diag(covar1))
            perr2 = np.sqrt(np.diag(covar2))
        for i in range(0,len(Edata)):
            w=Fis_Weight*Fissioning[Sys][i,1]/np.sum(Fissioning[Sys][:,1])
            fluxErr=Fissioning[Sys][i,2]
            if Edata[i]<=5.5:
                yA= Nagy(Edata[i], *params1)
                y[A] +=yA*w
                efit=(w*(montecarlo_error_propogation(Nagy, params1, perr1, Edata[i], False)[1]\
                   /Nagy(Edata[i], *params1)))**2
                err_fit[A]=err_fit[A]+efit
                err_flux[A]=err_flux[A]+(fluxErr*yA*w)**2
                err_data[A]=err_data[A]+(Uncert[i]*w)**2
            else: 
                yA= Nagy(Edata[i], *params2)
                y[A] +=yA*w
                efit=(w*(montecarlo_error_propogation(Nagy, params2, perr2, Edata[i], False)[1]\
                   /Nagy(Edata[i], *params1)))**2
                err_fit[A]=err_fit[A]+efit
                err_flux[A]=err_flux[A]+(fluxErr*yA*w)**2
                err_data[A]=err_data[A]+(Uncert[i]*w)**2
        err[A]=np.sqrt(err_fit[A]+err_flux[A]+err_data[A])
    return y,err

def GEF_FitDPLUS(U235_GEF,Fissioning,w):
    U235_fiss=Fissioning
    y={}
    err={}
    for k in U235_GEF.keys():
        yi=np.multiply((U235_fiss[:,1]/np.sum(U235_fiss[:,1])).reshape(-1,1),U235_GEF[k][:,0].reshape(-1,1)) # Total production of mass chains 
        err_data=np.square(np.multiply((U235_fiss[:,1]/np.sum(U235_fiss[:,1])).reshape(-1,1),U235_GEF[k][:,1].reshape(-1,1))) # Total error production of mass chains  
        err_flux=np.square(np.multiply(yi,U235_fiss[:,2].reshape(-1,1)))
        y[k]=w*np.sum(yi)
        err[k]=w*np.sqrt(np.add(np.sum(err_data),np.sum(err_flux)))
    return y,err

def ReadGEF(datapath):
    y={}
    j=-1
    k=0
    
    try: 
        ifile=open(datapath,'r')
        for line in ifile:
            if line[:7] == ' formed':
                split_list=line.split(' ')
                j+=1
    
            if line[:27] =='--- Mass-yield distribution' :
                for i in range(0,6):
                    line=ifile.next()
                while line[:22] != '        </Mass_yields>':
                    split_list=line.split()
    
                    if not split_list[0] in y.keys():  # this is the first time we've encountered this mass A
                    # let's initialize our dictionaries
                        y[split_list[0]] = np.zeros([46,2])
                        y[split_list[0]][:,1]=0.00001
                        y[split_list[0]][j,0] = float(split_list[2])
                        if float(split_list[2]) > 0:
                            y[split_list[0]][j,1] = float(split_list[3])
                        else:
                            y[split_list[0]][j,1] = 0.00001
                    else: # Mass chain seen before
                        y[split_list[0]][j,0] = float(split_list[2])
                        if float(split_list[2]) > 0:
                            y[split_list[0]][j,1]= float(split_list[3])
                        else:
                            y[split_list[0]][j,1] =  0.00001
                    line=ifile.next()
        # Close the file
        ifile.close()
    except IOError as e:
        print "I/O error({0}): {1}".format(e.errno, e.strerror)
        print "File not found was: {0}"
    return y


def Nagy(e, y, b):
    return y * np.exp(b * e)

def montecarlo_error_propogation(func, params, paramsErr, x, multivariate=False):
    if multivariate:
        sampledParams = np.random.multivariate_normal(params, paramsErr, 1000)
    else: 
        sampledParams = np.random.normal(params, paramsErr, (1000,len(params)))
    y=[]
    for p in sampledParams:
        y.append(func(x, *p))
    return np.mean(y), np.std(y) 

def ExpDataCheck(file,Fissioning,CWD): 
    
    # Reference Experimental Data 
    print 'Reading in experimental Data'
    U235_expData = openpyxl.load_workbook(CWD+'\\Data\\U235_Data.xlsx')  # 235 energy dependent data 
    U236_expData = openpyxl.load_workbook(CWD+'\\Data\\U236_Data.xlsx')  # 236 energy dependent data 
    U238_expData = openpyxl.load_workbook(CWD+'\\Data\\U238_Data.xlsx')  # 238 energy dependent data 
    Pu239_expData = openpyxl.load_workbook(CWD+'\\Data\\Pu239_Data.xlsx')  # 239 energy dependent data 
    Pu240_expData = openpyxl.load_workbook(CWD+'\\Data\\Pu240_Data.xlsx')  # 240 energy dependent data 
    
    SheetData=np.zeros(5,dtype='object')
    SheetData[0]=U235_expData.sheetnames
    SheetData[1]=U236_expData.sheetnames
    SheetData[2]=U238_expData.sheetnames
    SheetData[3]=Pu239_expData.sheetnames
    SheetData[4]=Pu240_expData.sheetnames
    for i in range(5): # Remove empty 
        if SheetData[i]==['Sheet1']: 
            SheetData[i]=[]
    
    # Downselect 
    ExpData={}
    TrackLoc=np.where(Fissioning['Track'] == 1)[0]
    # Gather needed data 
    for i in range(len(TrackLoc)):
        ExpData[TrackLoc[i]]=SheetData[TrackLoc[i]]
    Deleted=[]
    Keep=[]
    for i in range(len(TrackLoc)): 
        # Search over other files being checked
        Others=np.where(TrackLoc!=TrackLoc[i])[0]   
        for p in range(len(ExpData[TrackLoc[i]])): 
            k=ExpData[TrackLoc[i]][p]
            CheckOthers=1 # Switch to zero if one doesnt have it. 
            for j in Others: 
                if k not in ExpData[TrackLoc[j]]: 
                    Deleted.append(k)
                    CheckOthers=0
            if CheckOthers==1: 
                Keep.append(k)
    ListUse=[] # These are the fission product yields that all fissioning systems have data for
    for i in Keep:
           if i not in ListUse:
              ListUse.append(i)
    DeleteReport= []
    for i in Deleted:
           if i not in DeleteReport:
              DeleteReport.append(i)
    file.write('\n')
    file.write('Experimental Data Results\n')
    file.write("Isotope, Cumulative Yield [%] , Cumulative Yield Absolute Uncertainty [%]\n")

    if len(DeleteReport)>0.0: 
        file.write('Omitting data below because not available for all fission yields from input\n')
        for i in range(len(DeleteReport)):
            file.write("{}\n".format(DeleteReport[i]))
    
    print 'Deleted data for'
    print DeleteReport
    return file,ListUse
#---------------------------------------------------------------------------------------# 
def Build_Nagy_Weighted_FPs(List1,fname,e_bins,ff,weight,fis_err):    
    """
    Builds a fission product estimate using Nagy's fits. 
    
    Parameters
    ==========
    fname : string
        Path and name to the fission spectrum input file
    e_bins : list of floats
        Mid-point energies at which to calculate fission product production
    ff : float
        Fission fraction - fraction of the fissions for a given isotope        
    weight : list of floats
        A normalized fission spectrum distribution
    fis_err : list of floats 
        Relative error of fission spectrum distribution 
    
    Returns
    =======
    nagy_data : dictionary of lists
        A dictionary of lists for each A.  The list contains the weighted fission fragment production estimate for each energy.
    """
    
    y={}
    err={}
    b_vals=np.ndarray((0,3))   # Fitting parameter values 
    Yo_vals=np.ndarray((0,3))

    wb = openpyxl.load_workbook(fname)
    
    for tab in List1:
        # Import sheet object
        sheet = wb[tab]

        # Save isotope information
        split_sheet=sheet.title.split('_')
        A=int(split_sheet[1])
        Z=int(split_sheet[0])
        elem=split_sheet[2]
        #print A,Z,elem
        
        # Read in experimental data
        i=1
        exp_data=[]       # [E,y,err,reference]
        while sheet.cell(row=i, column=1).value != None:
            exp_data.append([sheet.cell(row=i, column=1).value, sheet.cell(row=i, column=2).value, \
                             sheet.cell(row=i, column=3).value,sheet.cell(row=i, column=4).value])
            i+=1
        exp_data=sorted(exp_data,key=lambda l:l[0])
        exp_data_e=np.asarray(exp_data)[:,0].astype(np.float)
        exp_data_y=np.asarray(exp_data)[:,1].astype(np.float)*100
        exp_data_err=np.asarray(exp_data)[:,2].astype(np.float)*100

        # Determine fitting parameters
        if A<=106 or A>=131:
            params1,covar1=curve_fit(Nagy, exp_data_e, exp_data_y,sigma=exp_data_err,absolute_sigma=True)
            params2=params1
            perr1 = np.sqrt(np.diag(covar1))
            perr2 = perr1
        else:
            ind=0
            while exp_data_e[ind]<=5.5:
                ind+=1
            params1,covar1=curve_fit(Nagy, exp_data_e[0:ind], exp_data_y[0:ind],sigma=exp_data_err[0:ind],absolute_sigma=True)
            params2,covar2=curve_fit(Nagy, exp_data_e[ind:], exp_data_y[ind:],sigma=exp_data_err[ind:],absolute_sigma=True)
            perr1 = np.sqrt(np.diag(covar1))
            perr2 = np.sqrt(np.diag(covar2))
        # Store fitting data 
        b_vals=np.append(b_vals,np.matrix([A,params1[1],perr2[1]]),axis=0)
        Yo_vals=np.append(Yo_vals,np.matrix([A,params1[0],perr2[0]]),axis=0)

        # Calculate Nagy data
        for i in range(0,len(e_bins)):
            if not A in y.keys():
                y[A] = 0
                err[A] = 0
            if e_bins[i]<=5.5:
                y[A] += Nagy(e_bins[i], *params1)*ff*weight[i]
                err[A] += (montecarlo_error_propogation(Nagy, params1, perr1, e_bins[i], False)[1]\
                                    /Nagy(e_bins[i], *params1))**2*ff*weight[i]
            else: 
                y[A] += (params2[0]*exp(params2[1]*e_bins[i]))*ff*weight[i]
                err[A] += ((montecarlo_error_propogation(Nagy, params2, perr2, e_bins[i], False)[1]\
                                    /Nagy(e_bins[i], *params1))**2*ff*weight[i] +(weight[i]*fis_err[i])**2)
    
    for A in y.keys():
        err[A]=sqrt(err[A])       
    return y, err, Yo_vals,b_vals
#---------------------------------------------------------------------------------------# 
def RunNagyFits(file,CWD,Fissioning,ListUse): 
    # Create inputs 

    void,bins=Echeck2(Fissioning,file)
    y = {}
    err = {}
    absErr = {}
    for i in range(len(ListUse)): 
        j= int(ListUse[i].split("_")[1])
        y[j]=0.0
        err[j]=0.0
        absErr[j]=0.0
    
    if Fissioning['Track'][0]==1:
        y_235, err_235,Yo_235,b_235=Build_Nagy_Weighted_FPs(ListUse,CWD+'\\Data\\U235_Data.xlsx',bins,Fissioning['Weights'][0],Fissioning['U235'][:,1]/np.sum(Fissioning['U235'][:,1]),Fissioning['U235'][:,2])
        for A in y_235.keys():
            y[A]=y[A]+y_235[A]
            err[A]=sqrt(err_235[A]**2+err[A]**2)
            absErr[A]=err[A]*y[A]
    if Fissioning['Track'][1]==1:
        y_236, err_236,Yo_236,b_236=Build_Nagy_Weighted_FPs(ListUse,CWD+'\\Data\\U236_Data.xlsx',bins,Fissioning['Weights'][1],Fissioning['U236'][:,1]/np.sum(Fissioning['U236'][:,1]),Fissioning['U236'][:,2])
        for A in y_236.keys():
            y[A]=y[A]+y_236[A]
            err[A]=sqrt(err_236[A]**2+err[A]**2)
            absErr[A]=err[A]*y[A]
    if Fissioning['Track'][2]==1:
        y_238, err_238,Yo_238,b_238=Build_Nagy_Weighted_FPs(ListUse,CWD+'\\Data\\U238_Data.xlsx',bins,Fissioning['Weights'][2],Fissioning['U238'][:,1]/np.sum(Fissioning['U238'][:,1]),Fissioning['U238'][:,2])
        for A in y_238.keys():
            y[A]=y[A]+y_238[A]
            err[A]=sqrt(err_238[A]**2+err[A]**2)
            absErr[A]=err[A]*y[A]
    if Fissioning['Track'][3]==1:
        y_239, err_239,Yo_239,b_239=Build_Nagy_Weighted_FPs(ListUse,CWD+'\\Data\\Pu239_Data.xlsx',bins,Fissioning['Weights'][3],Fissioning['Pu239'][:,1]/np.sum(Fissioning['Pu239'][:,1]),Fissioning['Pu239'][:,2])
        for A in y_239.keys():
            y[A]=y[A]+y_239[A]
            err[A]=sqrt(err_239[A]**2+err[A]**2)
            absErr[A]=err[A]*y[A]
    if Fissioning['Track'][4]==1:
        y_240, err_240,Yo_240,b_240=Build_Nagy_Weighted_FPs(ListUse,CWD+'\\Data\\Pu240_Data.xlsx',bins,Fissioning['Weights'][4],Fissioning['Pu240'][:,1]/np.sum(Fissioning['Pu240'][:,1]),Fissioning['Pu240'][:,2])
        for A in y_240.keys():
            y[A]=y[A]+y_240[A]
            err[A]=sqrt(err_240[A]**2+err[A]**2)
            absErr[A]=err[A]*y[A]
    return y,absErr


# ---------------------------Read input---------------------------------------
CWD=os.getcwd()
Files=['Pu239.txt','Pu240.txt','U235.txt','U238.txt','U236.txt']
idx=-1
for f in range(5): 
    idx=idx+1

    if os.path.isfile(CWD+'\\Inputs\\'+Files[idx])==False: 
        del Files[idx]
        idx=idx-1
Idx=1
while os.path.isfile(CWD+"\\Output\\"+str(datetime.datetime.fromtimestamp(time.time()).strftime('%Y_%m_%d'))+'_'+str(Idx)+".txt")==True:
    Idx=Idx+1
file = open(CWD+"\\Output\\"+str(datetime.datetime.fromtimestamp(time.time()).strftime('%Y_%m_%d'))+'_'+str(Idx)+".txt","w") 
file, Fissioning,Flag2,Track= GatherInput(CWD,file,Files)

#---------------------------Run GEF -----------------------------------------
file,y,err=run_GEF(Fissioning,file,Files,Flag2,CWD)
file.write('\n')
file.write("GEF Mass Chain Results\n")
file.write("Mass Chain [A] , Mass Chain Yield [%] , Mass Chain Yield Absolute Uncertainty [%]\n")
A = map(int, y.keys())
A=sorted(A)
A=map(str,A)
for i in A:
    file.write('{},{:04.3e},{:04.3e}\n'.format(i,y[i],err[i]))
file.write('\n')

#--------------------------Run Experimental Data -----------------------------
print("Performing fits of experimental data")
file,ListUse=ExpDataCheck(file,Fissioning,CWD)
y,absErr=RunNagyFits(file,CWD,Fissioning,ListUse)
# Combine the data into a single data set
A = map(int, y.keys())
A=sorted(A)
A=map(str,A)
for i in A:
    for j in range(len(ListUse)): 
        if int(ListUse[j].split('_')[1])==int(i): 
            file.write('{}-{},{:04.3e},{:04.3e}\n'.format(str(ListUse[j].split('_')[2]),str(ListUse[j].split('_')[1]),y[int(i)],absErr[int(i)]))
file.write('\n')

file.close()

