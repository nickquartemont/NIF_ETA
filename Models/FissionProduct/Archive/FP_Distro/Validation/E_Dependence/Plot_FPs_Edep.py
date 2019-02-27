#!/usr/bin/python

import sys    # Library to parse command-line arguments
import os
import openpyxl

import copy as cp
import numpy as np
import matplotlib.pyplot as plt

from matplotlib.patches import Rectangle
from scipy.optimize import curve_fit
from math import exp, sqrt
from scipy.stats import chisquare
from scipy import interpolate
from WahlSystematics_2002 import Wahl

def Nagy(e, y, b):
    return y * np.exp(b * e)

def montecarlo_error_propogation(func, params, paramsErr, x, multivariate=False):
    if multivariate:
        sampledParams = np.random.multivariate_normal(params, paramsErr, 1000)
    else: 
        sampledParams = np.random.normal(params, paramsErr, (1000,len(params)))
    y=[]
    for p in sampledParams:
        y.append(func(x, *p))
    return np.mean(y), np.std(y)  

# User Inputs
#---------------------------------------------------------------------------------------#  
ZAID=92235      # Target nucleus

rootdir=os.path.abspath(os.getcwd())+'/235/'

wb = openpyxl.load_workbook(rootdir+'235_data.xlsx')

# Optional plots
nagy=True
poly=False
spline=False#True
freya=True
wahl=True
#---------------------------------------------------------------------------------------#  

for tab in wb.get_sheet_names():
    # Import sheet object
    sheet = wb.get_sheet_by_name(tab)
    
    # Save isotope information
    split_sheet=sheet.title.split('_')
    A=int(split_sheet[1])
    Z=int(split_sheet[0])
    elem=split_sheet[2]
    print A,Z,elem
    
    # Read in experimental data
    i=1
    exp_data=[]       # [E,y,err,reference]
    while sheet.cell(row=i, column=1).value != None:
        exp_data.append([sheet.cell(row=i, column=1).value, sheet.cell(row=i, column=2).value, \
                         sheet.cell(row=i, column=3).value,sheet.cell(row=i, column=4).value])
        i+=1
    exp_data=sorted(exp_data,key=lambda l:l[0])
    exp_data_e=np.asarray(exp_data)[:,0].astype(np.float)
    exp_data_y=np.asarray(exp_data)[:,1].astype(np.float)*100
    exp_data_err=np.asarray(exp_data)[:,2].astype(np.float)*100
    
    # Read in GEF data
    gef_data=[]     # [E,y,err]
    tmp=0
    for subdir, dirs, files in os.walk(rootdir):
        if subdir != 'FREYA':
            for file in files:
                if file=="GEF.dat":
                    #print subdir, file
                    try: 
                        ifile = open(subdir+'/'+file, 'r') 
                        for line in ifile:
                            if line[:7] == ' formed':
                                E=float(line.split(' ')[8].strip())             
                                #print "Energy =", E

                            if line[:27] =='--- Mass-yield distribution':
                                for i in range(0,6):
                                    line=ifile.next()
                                while line[:22] != '        </Mass_yields>':
                                    split_list=line.split()
                                    if int(split_list[0].strip())==A:  
                                        gef_data.append([E,float(split_list[2]),float(split_list[3])])

                                    line=ifile.next()
                        # Close the file
                        ifile.close()
                    except IOError as e:
                        print "I/O error({0}): {1}".format(e.errno, e.strerror)
                        print "File not found was: {0}".format(fname)
    gef_data=sorted(gef_data,key=lambda l:l[0])
    gef_data_e=np.asarray(gef_data)[:,0].astype(np.float)
    gef_data_y=np.asarray(gef_data)[:,1].astype(np.float)
    gef_data_err=np.asarray(gef_data)[:,2].astype(np.float)
    
    # Calculate reduced chi-squared
    red_chi_gef=0.0
    for i in range(0,len(exp_data_e)):
        for j in range(0,len(gef_data_e)):
            if gef_data_e[j]==exp_data_e[i]:
                red_chi_gef+=(exp_data_y[i]-gef_data_y[j])**2 / exp_data_err[i]**2
    red_chi_gef=red_chi_gef/(len(exp_data_e)-1)
        
    # Calculate chisquare and p value
    f_obs=[]
    for i in range(0,len(exp_data_e)):
        for j in range(0,len(gef_data_e)):
            if gef_data_e[j]==exp_data_e[i]:
                f_obs.append(gef_data_y[j])
    (chi_gef,p_gef)=chisquare(f_obs, f_exp=exp_data_y)
    
    # Calculate Nagy fits: 
    if nagy:
        nagy_fit=[]
        nagy_fit_err=[]
        
        # Fitting function
        #def func(x, a, b):
        #    return a * np.exp(b * x)
        
        # Determine fitting parameters
        if A<=106 or A>=131:
            params1,covar1=curve_fit(Nagy, exp_data_e, exp_data_y,sigma=exp_data_err,absolute_sigma=True)
            params2=params1
            covar2=covar1
            perr1 = np.sqrt(np.diag(covar1))
            perr2=perr1
        else:
            ind=0
            while exp_data_e[ind]<=5.5:
                ind+=1
            params1,covar1=curve_fit(Nagy, exp_data_e[0:ind], exp_data_y[0:ind],sigma=exp_data_err[0:ind],absolute_sigma=True)
            params2,covar2=curve_fit(Nagy, exp_data_e[ind:], exp_data_y[ind:],sigma=exp_data_err[ind:],absolute_sigma=True)
            perr1 = np.sqrt(np.diag(covar1))
            perr2 = np.sqrt(np.diag(covar2))
        for e in exp_data_e:
            if e<=5.5:
                nagy_fit.append([e,params1[0]*exp(params1[1]*e)])
                nagy_fit_err.append(montecarlo_error_propogation(Nagy, params1, perr1, e, False)[1])
            else: 
                nagy_fit.append([e,params2[0]*exp(params2[1]*e)])
                nagy_fit_err.append(montecarlo_error_propogation(Nagy, params2, perr2, e, False)[1])
        nagy_fit=np.asarray(sorted(nagy_fit,key=lambda l:l[0]))
        nagy_fit_e=np.asarray(nagy_fit)[:,0].astype(np.float)
        nagy_fit_y=np.asarray(nagy_fit)[:,1].astype(np.float)
        nagy_fit_err=np.asarray(nagy_fit_err)
        
        # Calculate reduced chi-squared
        red_chi_nagy=0.0
        for i in range(0,len(nagy_fit_e)):
            red_chi_nagy+=(exp_data_y[i]-nagy_fit_y[i])**2 / exp_data_err[i]**2
        red_chi_nagy=red_chi_nagy/(len(nagy_fit_e)-1)
        
        # Calculate chisquare and p value
        (chi_nagy,p_nagy)=chisquare(nagy_fit_y, f_exp=exp_data_y)
       
    # Calculate Wahl sytematics: 
    if wahl:
        wahl_sys=[]
        
        # Calculate chain yields
        for e in exp_data_e:
            y={}
            for A_ch in range(60,180):
                y[A_ch]=Wahl(A_ch,ZAID,e)
            
            wahl_sys.append([e,y[A]*200/sum(y.values())])
        wahl_sys=np.asarray(sorted(wahl_sys,key=lambda l:l[0]))
        wahl_sys_e=np.asarray(wahl_sys)[:,0].astype(np.float)
        wahl_sys_y=np.asarray(wahl_sys)[:,1].astype(np.float)
        wahl_sys_err=wahl_sys_y*(25.*np.exp(-0.25*np.log2(wahl_sys_y)))/100
        
        # Calculate reduced chi-squared
        red_chi_wahl=0.0
        for i in range(0,len(wahl_sys_e)):
            red_chi_wahl+=(exp_data_y[i]-wahl_sys_y[i])**2 / exp_data_err[i]**2
        red_chi_wahl=red_chi_wahl/(len(wahl_sys_e)-1)
        
        # Calculate chisquare and p value
        (chi_wahl,p_wahl)=chisquare(wahl_sys_y, f_exp=exp_data_y)
        
    # Calculate Poly fits: 
    if poly:
        poly_fit=[]
        
        # Fitting function
        def func(x, a, b, c):
            return a * x**2 + b * x + c
        
        # Determine fitting parameters
        params,covar=curve_fit(func, exp_data_e, exp_data_y,sigma=exp_data_err)
    
        # Calculate Poly data
        for e in exp_data_e:
            poly_fit.append([e,func(e,params[0],params[1],params[2])])
        poly_fit=np.asarray(sorted(poly_fit,key=lambda l:l[0]))
        poly_fit_e=np.asarray(poly_fit)[:,0].astype(np.float)
        poly_fit_y=np.asarray(poly_fit)[:,1].astype(np.float)
        
        # Calculate reduced chi-squared
        red_chi_poly=0.0
        for i in range(0,len(exp_data_e)):
            for j in range(0,len(poly_fit_e)):
                if poly_fit_e[j]==exp_data_e[i]:
                    red_chi_poly+=(exp_data_y[i]-poly_fit_y[j])**2 / exp_data_err[i]**2
        red_chi_poly=red_chi_poly/(len(exp_data_e)-1)
        
        # Calculate chisquare and p value
        f_obs=[]
        for i in range(0,len(exp_data_e)):
            for j in range(0,len(poly_fit_e)):
                if poly_fit_e[j]==exp_data_e[i]:
                    f_obs.append(poly_fit_y[j])
        (chi_poly,p_poly)=chisquare(f_obs, f_exp=exp_data_y)
        
    # Calculate Spline fits: 
    if spline:
        spline_fit_e=exp_data_e
        spline_fit_y = interpolate.UnivariateSpline(exp_data_e, exp_data_y, s = 5e8)(exp_data_e)
        
        # Calculate reduced chi-squared
        red_chi_spline=0.0
        for i in range(0,len(spline_fit_e)):
            red_chi_spline+=(exp_data_y[i]-spline_fit_y[i])**2 / exp_data_err[i]**2
        red_chi_spline=red_chi_spline/(len(spline_fit_e)-1)
        
        # Calculate chisquare and p value
        (chi_spline,p_spline)=chisquare(spline_fit_y, f_exp=exp_data_y)
        
    # Read in FREYA data: 
    if freya:
        freya_data=[]
        for subdir, dirs, files in os.walk(rootdir+'FREYA/data'):
            for file in files:
                if file=="ff_yield.res":
                    #print subdir, file
                    E=float(subdir.split('\\')[-1])
                    try: 
                        ifile = open(subdir+'/'+file, 'r') 
                        for line in ifile:
                            split_list=line.split(':')
                            if int(split_list[0].strip())==A:  
                                freya_data.append([E,float(split_list[1]),float(split_list[2])])
                        # Close the file
                        ifile.close()
                    except IOError as e:
                        print "I/O error({0}): {1}".format(e.errno, e.strerror)
                        print "File not found was: {0}".format(fname)
        freya_data=sorted(freya_data,key=lambda l:l[0])
        freya_data_e=np.asarray(freya_data)[:,0].astype(np.float)
        freya_data_y=np.asarray(freya_data)[:,1].astype(np.float)*200
        freya_data_err=np.asarray(freya_data)[:,2].astype(np.float)*200
        
        # Calculate reduced chi-squared
        red_chi_freya=0.0
        for i in range(0,len(exp_data_e)):
            for j in range(0,len(freya_data_e)):
                if freya_data_e[j]==exp_data_e[i]:
                    #print exp_data_y[i],freya_data_y[j]
                    red_chi_freya+=(exp_data_y[i]-freya_data_y[j])**2 / exp_data_err[i]**2
        red_chi_freya=red_chi_freya/(len(exp_data_e)-1)
        
        # Calculate chisquare and p value
        f_obs=[]
        for i in range(0,len(exp_data_e)):
            for j in range(0,len(freya_data_e)):
                if freya_data_e[j]==exp_data_e[i]:
                    f_obs.append(freya_data_y[j])
        (chi_freya,p_freya)=chisquare(f_obs, f_exp=exp_data_y)
        
    ######## Create Plot #########
    # Allow use of Tex sybols
    plt.rc('text', usetex=True)

    # Preset data set format scheme
    s=18
    j=0
    linewidth=['1.5','3']
    marker=['*','o','v','^','<','>','1','2','3','4','8','s','p','*','h','H','+','x','d','D']
    linestyle=[':', '-.', '--', '-','-','-']
    dashes=[[2, 2, 2, 2], [10, 5, 2, 5], [10, 5, 10, 5], [10, 2, 2, 2, 2, 2],
            [10, 0.1], [10, 2, 10, 2, 2, 2, 2, 2]]

    # Set up figure
    fig = plt.figure(figsize=(12,8))
    ax1 = fig.add_axes([0.1, 0.1, 0.80, 0.85])

    # Set Line color cycle
    ax1.set_color_cycle(['k', 'k', 'k', 'k'])
    #ax1.set_color_cycle(['k','b','g','r','c','m','y','w'])
    
    # Set data point color options
    c=['k', 'k', 'k', 'k','k', 'k', 'k', 'k']

    # Set axes
#    ax1.axis([0, 15, 1E-2, 1E-1])
    #ax1.set_xscale('log')
    #ax1.set_yscale('log')

    # Set axes labels and plot title.
    #ax1.set_title('{}{}'.format(A,elem))    
    ax1.set_xlabel('\\textbf{Energy [MeV]}', fontsize=20)
    ax1.set_ylabel('\\textbf{Cumulative Yield}', fontsize=20)
    ax1.tick_params(axis='both', which='major', labelsize=20, width=2)
    ax1.tick_params(axis='both', which='minor', width=2)

    # Loop over the list of experimental data
    data_set=sorted(set(np.asarray(exp_data)[:,3]),reverse=False)
    lns=[]
    labs=[]
    hndls=[]
    for d in data_set:
        
        # Reset x and y data lists and index for lists
        x=[]
        y=[]
        err=[]

        # Build list of x, y coord pairs from file 
        for i in range(0,len(exp_data)): 
            if exp_data[i][3]==d:
                x.append(float(exp_data[i][0]))
                y.append(float(exp_data[i][1])*100)
                err.append(float(exp_data[i][2])*100)
        
        # Add data set to scatter plot
        x=np.float64(x) # Required for scatter plots 
        y=np.float64(y) # Required for scatter plots
        #ax1.scatter(x[:], y[:], s=s, c=c[0], marker=marker[j], label=d) 
        ax1.errorbar(x,y,err, markersize=6, fmt=marker[j],label=d)

        j+=1

    # Add model calculations to plot
    #ax1.plot(gef_data_e,gef_data_y,label="GEF-v2.2, $\chi_r$={:2.2f}, $\chi_p$={:2.2f}, p={:2.2f}".format(red_chi_gef, chi_gef,p_gef))
    ax1.plot(gef_data_e,gef_data_y,label="GEF-v2.2, $\chi_r$={:2.2f}".format(red_chi_gef, chi_gef,p_gef),ls='-', color='k', marker=None)
    
    # Define 1 sigm error bounds
    upper=gef_data_y+gef_data_err
    lower=gef_data_y-gef_data_err
    rect = Rectangle((0, 0), 1, 1, fc="grey", alpha=0.8)
    ax1.fill_between(gef_data_e, upper, lower, facecolor='grey', alpha=0.8)
    
    # Add FREYA data to plot
    if freya:
        #ax1.plot(freya_data_e,freya_data_y,label="FREYA v1.2, $\chi_r$={:2.2f}, $\chi_p$={:2.2f}, p={:2.2f}".format(red_chi_freya, chi_freya,p_freya),ls=':')
        ax1.plot(freya_data_e,freya_data_y,label="FREYA v1.2, $\chi_r$={:2.2f}".format(red_chi_freya, chi_freya,p_freya),ls=':', color='k', marker=None)
        # Define 1 sigm error bounds
        upper=freya_data_y+freya_data_err
        lower=freya_data_y-freya_data_err
        freya_rect = Rectangle((0, 0), 1, 1, fc="grey", alpha=0.6)
        ax1.fill_between(freya_data_e, upper, lower, facecolor='grey', alpha=0.6)
    
    # Add Nagy fits to plot
    if nagy:
        #ax1.plot(nagy_fit_e,nagy_fit_y,label="Nagy Fit, $\chi_r$={:2.2f}, $\chi_p$={:2.2f}, p={:2.2f}".format(red_chi_nagy, chi_nagy,p_nagy),ls='--')
        ax1.plot(nagy_fit_e,nagy_fit_y,label="Nagy Fit, $\chi_r$={:2.2f}".format(red_chi_nagy, chi_nagy,p_nagy),ls='--', color='k', marker=None)
        # Define 1 sigm error bounds
        upper=nagy_fit_y+nagy_fit_err
        lower=nagy_fit_y-nagy_fit_err
        nagy_rect = Rectangle((0, 0), 1, 1, fc="grey", alpha=0.4)
        ax1.fill_between(nagy_fit_e, upper, lower, facecolor='grey', alpha=0.4)
    
    # Add Wahl Systematics to plot
    if wahl:
        #ax1.plot(wahl_sys_e,wahl_sys_y,label="Wahl Sys, $\chi_r$={:2.2f}, $\chi_p$={:2.2f}, p={:2.2f}".format(red_chi_wahl, chi_wahl,p_wahl),ls='-.')
        ax1.plot(wahl_sys_e,wahl_sys_y,label="Wahl Sys, $\chi_r$={:2.2f}".format(red_chi_wahl, chi_wahl,p_wahl),ls='-.', color='k', marker=None)
        # Define 1 sigm error bounds
        upper=wahl_sys_y+wahl_sys_err
        lower=wahl_sys_y-wahl_sys_err
        wahl_rect = Rectangle((0, 0), 1, 1, fc="grey", alpha=0.15)
        ax1.fill_between(wahl_sys_e, upper, lower, facecolor='grey', alpha=0.15)
    
    # Add Poly fits to plot
    if poly:
        ax1.plot(poly_fit_e,poly_fit_y,label="Poly Fit, $\chi_r$={:2.2f}, $\chi_p$={:2.2f}, p={:2.2f}".format(red_chi_poly, chi_poly,p_poly),ls=':')
    
    # Add Spline fits to plot
    if spline:
        ax1.plot(spline_fit_e,spline_fit_y,label="Spline Fit, $\chi_r$={:2.2f}, $\chi_p$={:2.2f}, p={:2.2f}".format(red_chi_spline, chi_spline,p_spline),ls='-',dashes=[10, 2, 2, 2, 2, 2])
        
    # Add and locate legend
    leg = ax1.legend()
    handles, labels = ax1.get_legend_handles_labels() # get handles
#	handles = [h[0] for h in handles] # remove the errorbars
    handles.insert(1,rect)
    #labels.insert(1,'\\textbf{GEF 1\\boldmath{$\sigma$} Range}')
    labels.insert(1,'GEF 1$\sigma$ Range')
    if freya:
        handles.insert(3,freya_rect)
        labels.insert(3,'FREYA 1$\sigma$ Range')
    if Nagy:
        handles.insert(5,nagy_rect)
        labels.insert(5,'Nagy 1$\sigma$ Range')
    if Wahl:
        handles.insert(7,wahl_rect)
        labels.insert(7,'Wahl 1$\sigma$ Range')
    ax1.legend(handles, labels, borderaxespad=0.75, loc='center left', bbox_to_anchor=(1,0.5), fontsize=14, 
             handlelength=2, borderpad=0.5, labelspacing=0.75, fancybox=True, 
             framealpha=1.0, numpoints = 1);

    #plt.show()
    fig.savefig(rootdir+'{}_{}_{}_lin.png'.format(Z,A,elem), bbox_inches='tight')
    ax1.set_yscale('log')
    fig.savefig(rootdir+'{}_{}_{}_log.png'.format(Z,A,elem), bbox_inches='tight')