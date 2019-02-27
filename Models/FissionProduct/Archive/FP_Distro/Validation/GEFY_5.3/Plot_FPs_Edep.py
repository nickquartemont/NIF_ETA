#!/usr/bin/python

import sys    # Library to parse command-line arguments
import os
import csv
import openpyxl
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.patches import Rectangle

rootdir=os.path.abspath(os.getcwd())+'/238U/'
gefdir=os.path.abspath(os.getcwd())+'/238U/ChainYields/'
expdir=os.path.join(os.path.abspath(os.getcwd()), os.pardir)+'/E_Dependence/238/'

wb = openpyxl.load_workbook(expdir+'238_data.xlsx')

for tab in wb.get_sheet_names():
    # Import sheet object
    sheet = wb.get_sheet_by_name(tab)
    
    # Save isotope information
    split_sheet=sheet.title.split('_')
    A=int(split_sheet[1])
    Z=int(split_sheet[0])
    elem=split_sheet[2]
    print A,Z,elem
    
    # Read in experimental data
    i=1
    exp_data=[]       # [E,y,err,reference]
    while sheet.cell(row=i, column=1).value != None:
        exp_data.append([sheet.cell(row=i, column=1).value, sheet.cell(row=i, column=2).value, \
                         sheet.cell(row=i, column=3).value,sheet.cell(row=i, column=4).value])
        i+=1
    
    # Read in GEF data
    gef_data=[]     # [E,y,err]
    for subdir, dirs, files in os.walk(gefdir):
        for file in files:
            print subdir, file
            E=float(file.split('_')[0].strip())/1E6          
            print "Energy =", E
            try: 
                ifile = open(subdir+'/'+file, 'r') 
                reader = csv.reader(ifile)
                for line in reader:
                    if int(line[1].strip())==A: 
                        gef_data.append([E,float(line[2]),float(line[3])])
                # Close the file
                ifile.close()
            except IOError as e:
                print "I/O error({0}): {1}".format(e.errno, e.strerror)
                print "File not found was: {0}".format(fname)
    gef_data=sorted(gef_data,key=lambda l:l[0])
    print gef_data
    
    ######## Create Plot #########
    # Allow use of Tex sybols
    plt.rc('text', usetex=True)

    # Preset data set format scheme
    s=18
    j=0
    linewidth=['1.5','3']
    marker=['*','o','v','^','<','>','1','2','3','4','8','s','p','*','h','H','+','x','d','D']
    linestyle=[':', '-.', '--', '-','-','-']
    dashes=[[2, 2, 2, 2], [10, 5, 2, 5], [10, 5, 10, 5], [10, 2, 2, 2, 2, 2],
            [10, 0.1], [10, 2, 10, 2, 2, 2, 2, 2]]

    # Set up figure
    fig = plt.figure()
    #ax1 = fig.add_subplot(111)
    ax1 = fig.add_axes([0.1, 0.1, 0.80, 0.85])

    # Set Line color cycle
    ax1.set_color_cycle(['k', 'k', 'k', 'k'])
    #ax1.set_color_cycle(['k','b','g','r','c','m','y','w'])
    
    # Set data point color options
    c=['k', 'k', 'k', 'k','k', 'k', 'k', 'k']

    # Set axes
#    ax1.axis([0, 15, 1E-2, 1E-1])
    #ax1.set_xscale('log')
    #ax1.set_yscale('log')

    # Set axes labels and plot title.
    ax1.set_title('{}{}'.format(A,elem))    
    ax1.set_xlabel('\\textbf{Energy [MeV]}', fontsize=14)
    ax1.set_ylabel('\\textbf{Cumulative Yield}', fontsize=14)

    # Loop over the list of experimental data
    data_set=sorted(set(np.asarray(exp_data)[:,3]),reverse=False)
    lns=[]
    labs=[]
    hndls=[]
    for d in data_set:
        
        # Reset x and y data lists and index for lists
        x=[]
        y=[]
        err=[]

        # Build list of x, y coord pairs from file 
        for i in range(0,len(exp_data)): 
            if exp_data[i][3]==d:
                x.append(float(exp_data[i][0]))
                y.append(float(exp_data[i][1])*100)
                err.append(float(exp_data[i][2])*100)
        
        # Add data set to scatter plot
        x=np.float64(x) # Required for scatter plots 
        y=np.float64(y) # Required for scatter plots
        #ax1.scatter(x[:], y[:], s=s, c=c[0], marker=marker[j], label=d) 
        ax1.errorbar(x,y,err, markersize=6, fmt=marker[j],label=d)
	#lns.append(ax1.errorbar(x,y,err, markersize=6, fmt=marker[j],label=d))
	#labs.append(lns[j].get_label())
	#handles, labels = ax1.get_legend_handles_labels() # get handles
	#hndls.append(handles[j][0])

        j+=1
    
    # Build list of x, y coord pairs from model calculations
    x=[]
    y=[]
    err=[] 
    for i in range(0,len(gef_data)): 
        x.append(float(gef_data[i][0]))
        y.append(float(gef_data[i][1])*100)
        err.append(float(gef_data[i][2])*100)

    # Add model calculations to plot
    ax1.plot(x,y,label="GEF-v2.2")
    
    # Define 1 sigm error bounds
    y=np.asarray(y)
    err=np.asarray(err)
    upper=y+err
    lower=y-err
    rect = Rectangle((0, 0), 1, 1, fc="grey", alpha=0.5)
    ax1.fill_between(x, upper, lower, facecolor='grey', alpha=0.5,
                label='1 sigma range')
    
    # Add and locate legend
    leg = ax1.legend()
    handles, labels = ax1.get_legend_handles_labels() # get handles
#	handles = [h[0] for h in handles] # remove the errorbars
    handles.insert(1,rect)
    #labels.insert(1,'\\textbf{GEF 1\\boldmath{$\sigma$} Range}')
    labels.insert(1,'GEF 1$\sigma$ Range')
    ax1.legend(handles, labels, borderaxespad=0.75, loc=0, fontsize=12, 
             handlelength=2, borderpad=0.5, labelspacing=0.75, fancybox=True, 
             framealpha=1.0, numpoints = 1);

    #plt.show()
    fig.savefig(rootdir+'Figs/{}_{}_{}_lin.png'.format(Z,A,elem), bbox_inches='tight')

    ax1.set_yscale('log')
    fig.savefig(rootdir+'Figs/{}_{}_{}_log.png'.format(Z,A,elem), bbox_inches='tight')
